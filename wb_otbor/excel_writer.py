"""Запись итогового файла Отбор на основе xlsx-шаблона."""
from copy import copy
from datetime import datetime
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from . import config
from . import settings as app_settings
from .logging_setup import get_logger


logger = get_logger('excel_writer')


def _append_applied_filters_sheet(wb, end_date, output_date: str):
    """
    Добавляет второй лист 'Применённые фильтры' с текущими настройками
    (фильтры справочника + числовые параметры + даты периода).
    """
    s = app_settings.load()

    # Удаляем старый лист, если есть (при пересохранении)
    for existing in list(wb.sheetnames):
        if existing == 'Применённые фильтры':
            del wb[existing]

    ws = wb.create_sheet('Применённые фильтры')

    thin = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill('solid', fgColor='1F4E78')
    header_font = Font(bold=True, color='FFFFFF', name='Calibri')
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='top', wrap_text=True)

    def _write_header(row, title):
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.cell(row=row, column=2).fill = header_fill
        ws.cell(row=row, column=2).border = border
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

    # --- Шапка с датой формирования ---
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ws['A1'] = 'Параметры отчёта WB'
    ws['A1'].font = Font(bold=True, size=14, color='1F4E78')
    ws.merge_cells('A1:B1')
    ws['A2'] = f'Дата формирования:'
    ws['B2'] = now

    # --- Период отчёта ---
    period_days = s.get('period_days', 7)
    offset_days = s.get('offset_from_today', 1)
    from datetime import timedelta
    start_date = end_date - timedelta(days=period_days - 1)
    ws['A3'] = 'Период:'
    ws['B3'] = f'{start_date.strftime("%d.%m.%Y")} — {end_date.strftime("%d.%m.%Y")}'
    ws['A4'] = 'Длина периода, дней:'
    ws['B4'] = period_days
    ws['A5'] = 'Сдвиг от сегодня, дней:'
    ws['B5'] = offset_days

    # --- Пороги Технички ---
    _write_header(7, 'Пороги «Технички» / «Отбора»')
    rows_th = [
        ('Мин. остаток (шт.):', s.get('min_stock', 10)),
        ('Мин. дистрибуция (%):', s.get('min_distrib_percent', 20)),
        ('Мин. показы:', s.get('min_shows', 1200)),
    ]
    for i, (lbl, val) in enumerate(rows_th, start=8):
        ws.cell(row=i, column=1, value=lbl)
        ws.cell(row=i, column=2, value=val)

    # --- Фильтры справочника ---
    start_row = 12
    _write_header(start_row, 'Фильтры справочника')
    row = start_row + 1

    filters = s.get('filters', {})
    for key in app_settings.FILTER_KEYS:
        values = filters.get(key, [])
        ws.cell(row=row, column=1, value=key).font = Font(bold=True)
        if values:
            val_text = '\n'.join(values) if len(values) <= 30 else \
                       '\n'.join(values[:30]) + f'\n…ещё {len(values) - 30}'
            ws.cell(row=row, column=2, value=val_text)
        else:
            ws.cell(row=row, column=2, value='(все — фильтр не задан)').font = \
                Font(italic=True, color='999999')
        ws.cell(row=row, column=2).alignment = left
        row += 1

    # Ширина колонок
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 80

    # Заморозим шапку
    ws.freeze_panes = 'A2'

    logger.info("Второй лист 'Применённые фильтры' добавлен.")


def write_otbor_file(df_base: pd.DataFrame, end_date, template_path: Path,
                     output_path: Path, log=print) -> Path:
    """
    Заполняет шаблон данными из df_base и сохраняет под output_path.
    Второй лист 'Применённые фильтры' с текущими настройками добавляется автоматически.
    """
    logger.info(f"write_otbor_file: template={template_path}, output={output_path}, "
                f"rows={len(df_base)}")

    # --- Проверка входных данных ---
    if not Path(template_path).exists():
        logger.error(f"Шаблон не найден: {template_path}")
        raise FileNotFoundError(f"Шаблон не найден: {template_path}")

    required_cols = {'Артикул', 'Показы', 'Клики', 'Заказы',
                     'Количество фото (-2 от скрипта)',
                     'ФИО менеджера',
                     'Дата создания на WB'}
    missing = required_cols - set(df_base.columns)
    if missing:
        logger.error(f"В df_base отсутствуют столбцы: {missing}")
        raise ValueError(f"В df_base отсутствуют обязательные столбцы: {missing}")

    try:
        wb = load_workbook(template_path)
        ws = wb.active
    except Exception:
        logger.exception(f"Не удалось открыть шаблон {template_path}")
        raise

    output_date = end_date.strftime('%d.%m')

    # ===== НОВАЯ РАСКЛАДКА КОЛОНОК (после вставки 'ФИО менеджера' на позицию I) =====
    #  A=1  Артикул
    #  B=2  Артикул WB
    #  C=3  Бизнес-группа
    #  D=4  Розничный отдел
    #  E=5  Группа
    #  F=6  Сезон
    #  G=7  Бренд
    #  H=8  Ответственный за группу
    #  I=9  ФИО менеджера         ← новая
    #  J=10 Коллекция
    #  K=11 Показы
    #  L=12 Клики
    #  M=13 Заказы
    #  N=14 CTR          (формула)
    #  O=15 Конверсия    (формула)
    #  P=16 Остаток
    #  Q=17 Дистрибуция
    #  R=18 Техничка
    #  S=19 Отбор для поиска        (формула / текст)
    #  T=20 Отбор по CTR            (формула)
    #  U=21 Отбор по CR             (формула)
    #  V=22 Отбор для поиска (вг)   (формула)
    #  W=23 Отбор по CTR (вг)       (формула)
    #  X=24 Отбор по CR  (вг)       (формула)
    #  Y=25 Количество фото (-2)
    #  Z=26 Дата создания на WB         (формат d.m.YYYY, текст)
    # =================================================================================
    TOTAL_COLS = 26

    # Сохраняем стили из шаблонной строки 3
    row3_styles = {}
    for col in range(1, TOTAL_COLS + 1):
        cell = ws.cell(row=3, column=col)
        row3_styles[col] = {
            'font': copy(cell.font),
            'fill': copy(cell.fill),
            'border': copy(cell.border),
            'alignment': copy(cell.alignment),
            'number_format': cell.number_format,
        }

    # Очищаем все строки данных
    for row in range(3, ws.max_row + 1):
        for col in range(1, TOTAL_COLS + 1):
            ws.cell(row=row, column=col).value = None

    # Имена столбцов из df_base
    остаток_col = [c for c in df_base.columns if 'Остаток' in c][0]
    distrib_col = [c for c in df_base.columns if 'Дистрибуция' in c][0]

    # Пороги для Технички/Отбора — читаем из settings.json
    min_stock = app_settings.get_numeric('min_stock')
    min_distrib = app_settings.get_numeric('min_distrib_percent')
    min_shows = app_settings.get_numeric('min_shows')

    last_data_row = 2 + len(df_base)

    log(f"Запись {len(df_base)} строк в Excel...")

    for idx, row_data in df_base.reset_index(drop=True).iterrows():
        r = 3 + idx

        # A–J: справочные данные
        ws.cell(row=r, column=1).value = str(row_data['Артикул'])
        ws.cell(row=r, column=2).value = (
            str(int(float(row_data['Артикул WB'])))
            if pd.notna(row_data['Артикул WB']) else ''
        )
        ws.cell(row=r, column=3).value = row_data['Бизнес-группа']
        ws.cell(row=r, column=4).value = row_data['Розничный отдел']
        ws.cell(row=r, column=5).value = row_data['Группа']
        ws.cell(row=r, column=6).value = row_data['Сезон']
        ws.cell(row=r, column=7).value = row_data['Бренд']
        ws.cell(row=r, column=8).value = row_data['Ответственный за группу']
        # I=9: ФИО менеджера (новая колонка из Распределение категорий.xlsx)
        ws.cell(row=r, column=9).value = row_data.get('ФИО менеджера', 'Не определено')
        ws.cell(row=r, column=10).value = row_data['Коллекция']

        # K–M: маркетинг
        ws.cell(row=r, column=11).value = int(row_data['Показы'])
        ws.cell(row=r, column=12).value = int(row_data['Клики'])
        ws.cell(row=r, column=13).value = int(row_data['Заказы'])

        # N, O: формулы CTR и Конверсии (было M, N → сдвиг на +1)
        ws.cell(row=r, column=14).value = f'=IFERROR(L{r}/K{r},0)'
        ws.cell(row=r, column=15).value = f'=IFERROR(M{r}/L{r},0)'

        # P: Остаток
        ws.cell(row=r, column=16).value = int(row_data[остаток_col])

        # Q: Дистрибуция (как decimal для формата 0.0%)
        ws.cell(row=r, column=17).value = row_data[distrib_col] / 100

        # R и S: Техничка и Отбор для поиска
        stock = row_data[остаток_col]
        distrib = row_data[distrib_col]
        shows = row_data['Показы']

        if stock < min_stock or distrib < min_distrib:
            ws.cell(row=r, column=19).value = 'Мало остатка'
            ws.cell(row=r, column=18).value = 0
        elif shows < min_shows:
            ws.cell(row=r, column=19).value = 'Мало показов'
            ws.cell(row=r, column=18).value = 0
        else:
            # $S$1 → $T$1, $T$1 → $U$1; M3 → N3, N3 → O3
            ws.cell(row=r, column=19).value = (
                f'=IF(OR(N{r}<$T$1*0.5,O{r}<$U$1*0.5),'
                f'"код для проверки","нормальный код")'
            )
            ws.cell(row=r, column=18).value = 1

        # T–X: формулы отбора (было S–W, сдвиг +1)
        # Общий сдвиг в ссылках: J→K, K→L, L→M, M→N, N→O, Q→R
        ws.cell(row=r, column=20).value = (
            f'=IF(N{r}<$T$1*0.5,"Низшая четверть",'
            f'IF(N{r}<$T$1,"Вторая четверть","Больше половины"))'
        )
        ws.cell(row=r, column=21).value = (
            f'=IF(O{r}<$U$1*0.5,"Низшая четверть",'
            f'IF(O{r}<$U$1,"Вторая четверть","Больше половины"))'
        )
        ws.cell(row=r, column=22).value = (
            f'=IF(OR(N{r}<SUMIFS(L:L,R:R,1,E:E,E{r})/SUMIFS(K:K,R:R,1,E:E,E{r})*0.5,'
            f'O{r}<SUMIFS(M:M,R:R,1,E:E,E{r})/SUMIFS(L:L,R:R,1,E:E,E{r})*0.5),'
            f'"код для проверки","нормальный код")'
        )
        ws.cell(row=r, column=23).value = (
            f'=IF(N{r}<SUMIFS(L:L,R:R,1,E:E,E{r})/SUMIFS(K:K,R:R,1,E:E,E{r})*0.5,'
            f'"Низшая четверть",IF(N{r}<SUMIFS(L:L,R:R,1,E:E,E{r})/SUMIFS(K:K,R:R,1,E:E,E{r}),'
            f'"Вторая четверть","Больше половины"))'
        )
        ws.cell(row=r, column=24).value = (
            f'=IF(O{r}<SUMIFS(M:M,R:R,1,E:E,E{r})/SUMIFS(L:L,R:R,1,E:E,E{r})*0.5,'
            f'"Низшая четверть",IF(O{r}<SUMIFS(M:M,R:R,1,E:E,E{r})/SUMIFS(L:L,R:R,1,E:E,E{r}),'
            f'"Вторая четверть","Больше половины"))'
        )

        # Y: Количество фото (-2)
        ws.cell(row=r, column=25).value = int(row_data['Количество фото (-2 от скрипта)'])

        # Z: Дата создания на WB (строка вида '25.4.2026' или пусто)
        ws.cell(row=r, column=26).value = str(row_data.get('Дата создания на WB', '') or '')

        # Применяем стили из шаблонной строки 3
        for col in range(1, TOTAL_COLS + 1):
            cell = ws.cell(row=r, column=col)
            style = row3_styles.get(col)
            if style:
                cell.font = copy(style['font'])
                cell.fill = copy(style['fill'])
                cell.border = copy(style['border'])
                cell.alignment = copy(style['alignment'])
                cell.number_format = style['number_format']

    # Обновляем формулы строки 1 под новый диапазон (сдвиг J→K, K→L, L→M, M→N, N→O, S→T, T→U)
    ws['K1'] = f'=SUBTOTAL(9,K3:K{last_data_row})'    # Показы
    ws['L1'] = f'=SUBTOTAL(9,L3:L{last_data_row})'    # Клики
    ws['M1'] = f'=SUBTOTAL(9,M3:M{last_data_row})'    # Заказы
    ws['N1'] = '=L1/K1'                                # CTR
    ws['O1'] = '=M1/L1'                                # Конверсия
    ws['T1'] = '=SUMIF(R:R,1,L:L)/SUMIF(R:R,1,K:K)'   # CTR Технички
    ws['U1'] = '=SUMIF(R:R,1,M:M)/SUMIF(R:R,1,L:L)'   # CR Технички

    # Обновляем заголовок столбца Остаток (теперь колонка P=16)
    ws.cell(row=2, column=16).value = f'Остаток на {output_date}'

    # Добавляем второй лист с применёнными фильтрами
    try:
        _append_applied_filters_sheet(wb, end_date, output_date)
    except Exception as exc:
        log(f"⚠ Не удалось добавить лист с фильтрами: {exc}")
        logger.exception("Ошибка при добавлении листа 'Применённые фильтры'")
        # Не падаем — основной лист важнее

    try:
        wb.save(output_path)
        logger.info(f"Excel-файл сохранён: {output_path}")
        log(f"Файл сохранён: {output_path}")
    except Exception:
        logger.exception(f"Не удалось сохранить {output_path}")
        raise
    return output_path
