"""
Загрузка маппинга «Группа → ФИО менеджера» из файла
«Распределение категорий.xlsx» в корне проекта.

Файл должен содержать лист с двумя колонками:
  - «Розничная группа» (например, «072 Трусы женские»)
  - «Ответственный»     (ФИО менеджера)

Если файл отсутствует, повреждён или группа не найдена —
возвращается «Не определено».
"""
from __future__ import annotations

import os
import unicodedata
from pathlib import Path
from openpyxl import load_workbook

from . import config
from .logging_setup import get_logger


logger = get_logger('manager_mapping')

MAPPING_FILENAME = 'Распределение категорий.xlsx'
MAPPING_FILE = config.BASE_DIR / MAPPING_FILENAME
UNKNOWN = 'Не определено'


def _nfc(s: str) -> str:
    """Нормализация Unicode в NFC (для сравнения строк на Windows/macOS)."""
    return unicodedata.normalize('NFC', s or '')


def _resolve_mapping_path() -> Path | None:
    """
    Ищет файл через os.listdir (надёжнее, чем Path.exists() на UNC+Cyrillic).
    Имя файла на диске может быть в NFD-форме (macOS/SMB-share), а литерал
    в коде — в NFC. Поэтому сравниваем нормализованные варианты.
    """
    try:
        folder = str(config.BASE_DIR)
        target = _nfc(MAPPING_FILENAME).lower()
        for name in os.listdir(folder):
            if _nfc(name).lower() == target:
                return Path(os.path.join(folder, name))
    except Exception:
        logger.exception(f"Не удалось прочитать содержимое {config.BASE_DIR}")
    return None

# Ключи-заголовки, которые ищем в файле (tolerant к регистру/пробелам).
KEY_COLUMN_NAMES = ('розничная группа', 'группа', 'retail group')
VALUE_COLUMN_NAMES = ('ответственный', 'фио', 'менеджер', 'manager')


def _normalize_key(value: str) -> str:
    """Нормализует название группы: NFC + trim + lower, чтобы matching был устойчив."""
    return _nfc(str(value or '').strip()).lower()


def load_mapping(path: Path = None) -> dict[str, str]:
    """
    Возвращает dict {нормализованная_группа: ФИО}.
    При ошибках — пустой dict и warning в лог.
    """
    if path is None:
        path = _resolve_mapping_path()

    if path is None or not os.path.exists(str(path)):
        logger.warning(f"Файл маппинга не найден: {MAPPING_FILE}. "
                       f"Все строки получат '{UNKNOWN}'.")
        return {}

    try:
        wb = load_workbook(str(path), data_only=True)
        ws = wb.active

        # Находим индексы колонок по заголовку (строка 1)
        headers = {}
        for col_idx in range(1, ws.max_column + 1):
            h = ws.cell(row=1, column=col_idx).value
            if h is None:
                continue
            headers[_normalize_key(h)] = col_idx

        key_col = None
        for cand in KEY_COLUMN_NAMES:
            if cand in headers:
                key_col = headers[cand]
                break
        val_col = None
        for cand in VALUE_COLUMN_NAMES:
            if cand in headers:
                val_col = headers[cand]
                break

        if not key_col or not val_col:
            logger.error(
                f"В {path.name} не найдены нужные колонки. "
                f"Ожидались: 'Розничная группа' и 'Ответственный'. "
                f"Найдены: {list(headers.keys())}"
            )
            return {}

        mapping: dict[str, str] = {}
        for r in range(2, ws.max_row + 1):
            k = ws.cell(row=r, column=key_col).value
            v = ws.cell(row=r, column=val_col).value
            if k is None or v is None:
                continue
            nk = _normalize_key(k)
            if not nk:
                continue
            mapping[nk] = str(v).strip()

        logger.info(f"Загружен маппинг 'Группа → ФИО менеджера': "
                    f"{len(mapping)} записей из {path.name}")
        return mapping
    except Exception:
        logger.exception(f"Ошибка чтения файла маппинга {path}")
        return {}


def get_manager(group_name: str, mapping: dict[str, str]) -> str:
    """
    Возвращает ФИО менеджера по названию группы.
    Если в маппинге группы нет — «Не определено».
    """
    if not group_name:
        return UNKNOWN
    return mapping.get(_normalize_key(group_name), UNKNOWN)
