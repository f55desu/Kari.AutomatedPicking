"""Работа с WB Content API + xlsx-кэш данных по карточкам.

Кэш хранит для каждого nmID:
  - photos (int)        — количество фото на карточке
  - created_at (str)    — ISO-дата создания карточки на WB (или пустая строка)

Старый формат кэша (только nmID + photos) поддерживается на чтение —
created_at просто будет пустым.
"""
import os
import time
from pathlib import Path
from openpyxl import Workbook, load_workbook

from . import config
from ._wb_content import (
    load_env_file,
    get_api_token,
    wb_post_json,
    count_photos,
    REQUEST_DELAY_SECONDS,
)
from .logging_setup import get_logger


logger = get_logger('wb_api')


# --- Структура записи в кэше ---
def _empty_record() -> dict:
    return {'photos': 0, 'created_at': ''}


# ---------------------------------------------------------------------
# Сохранение / загрузка xlsx-кэша
# ---------------------------------------------------------------------

def save_photo_cache(records: dict, cache_path: Path = None) -> Path:
    """
    Сохраняет dict {str(nmID) -> {'photos': int, 'created_at': str}} в xlsx.
    Возвращает путь.
    """
    cache_path = Path(cache_path) if cache_path else config.PHOTO_CACHE_FILE
    try:
        cache_path.parent.mkdir(parents=True, exist_ok=True)

        wb_cache = Workbook()
        ws = wb_cache.active
        ws.title = "photo_counts"
        ws.append(["Артикул WB", "Количество фото", "Дата создания"])

        def _sort_key(pair):
            k = pair[0]
            return int(k) if k.isdigit() else 0

        for nm_id, rec in sorted(records.items(), key=_sort_key):
            ws.append([
                nm_id,
                int(rec.get('photos', 0)),
                rec.get('created_at', '') or '',
            ])
        wb_cache.save(cache_path)
        logger.info(
            f"Фото-кэш сохранён: {cache_path.resolve()} ({len(records)} записей)"
        )
        return cache_path.resolve()
    except Exception:
        logger.exception(f"Не удалось сохранить фото-кэш в {cache_path}")
        raise


def load_photo_cache(cache_path: Path = None) -> dict:
    """
    Загружает dict {str(nmID) -> {'photos': int, 'created_at': str}} из xlsx.
    Поддерживает старый формат (без столбца "Дата создания").
    """
    cache_path = cache_path or config.PHOTO_CACHE_FILE
    if not cache_path.exists():
        logger.error(f"Фото-кэш не найден: {cache_path}")
        raise FileNotFoundError(f"Кэш не найден: {cache_path}")
    try:
        wb_cache = load_workbook(cache_path)
        ws = wb_cache.active
        result = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            nm_id = str(row[0]).strip()
            count = int(row[1]) if (len(row) > 1 and row[1] is not None) else 0
            created = ''
            if len(row) > 2 and row[2] is not None:
                created = str(row[2]).strip()
            result[nm_id] = {'photos': count, 'created_at': created}
        logger.info(f"Фото-кэш загружен: {cache_path} ({len(result)} записей)")
        return result
    except Exception:
        logger.exception(f"Ошибка при чтении фото-кэша {cache_path}")
        raise


# ---------------------------------------------------------------------
# Запрос к WB Content API
# ---------------------------------------------------------------------

def fetch_photos_from_api(target_nm_ids, timeout: int = 60, log=print) -> dict:
    """
    Batch-загрузка через cursor-пагинацию /content/v2/get/cards/list.

    Возвращает {str(nmID) -> {'photos': int, 'created_at': str}}.

    WB Content API v2 НЕ поддерживает массив nmID в одном запросе.
    Cursor-пагинация по всему каталогу — максимально быстрый легитимный метод.
    """
    load_env_file(config.ENV_FILE)
    token = get_api_token()

    target_set = set()
    for x in target_nm_ids:
        if x is None:
            continue
        s = str(x).strip()
        if not s or s.lower() == 'nan':
            continue
        try:
            target_set.add(str(int(float(s))))
        except (TypeError, ValueError):
            continue

    results: dict[str, dict] = {}
    cursor_payload = {"limit": 100}
    total_fetched = 0
    page = 0
    _first_card_dumped = False  # debug-дамп первой карточки в лог

    log(f"Целевых артикулов WB: {len(target_set)}")
    start = time.time()

    while True:
        page += 1
        payload = {"settings": {"cursor": cursor_payload, "filter": {"withPhoto": -1}}}

        data = wb_post_json(token, payload, timeout)
        cards = data.get("cards", [])
        if not cards:
            break

        # === DEBUG: один раз дампим все поля первой карточки ===
        # Это даёт возможность убедиться, какие именно ключи возвращает WB API.
        if not _first_card_dumped and cards:
            first = cards[0]
            top_keys = list(first.keys())
            logger.info(f"DEBUG: WB API вернул карточку с полями верхнего уровня: {top_keys}")
            # Полный дамп первой карточки в DEBUG (попадёт в logs/wb_otbor.log)
            import json as _json
            try:
                logger.debug(f"DEBUG full card[0]:\n{_json.dumps(first, ensure_ascii=False, indent=2)[:4000]}")
            except Exception:
                logger.debug(f"DEBUG card[0] (не сериализуется): {first!r}"[:2000])
            # Подсветка полей с датами
            date_like = {k: first[k] for k in first
                         if 'created' in k.lower() or 'updated' in k.lower()
                         or k.lower() in ('date',)}
            if date_like:
                log(f"DEBUG: поля-даты в первой карточке: {date_like}")
                logger.info(f"DEBUG: поля-даты: {date_like}")
            else:
                log(f"DEBUG: в первой карточке НЕТ полей с датами!")
                logger.warning(f"DEBUG: в первой карточке НЕТ полей с датами. "
                               f"Полный список ключей: {top_keys}")
            _first_card_dumped = True

        for card in cards:
            nm_id_str = str(card.get("nmID", ""))
            if nm_id_str in target_set:
                results[nm_id_str] = {
                    'photos':     count_photos(card),
                    'created_at': str(card.get('createdAt', '') or ''),
                }

        total_fetched += len(cards)
        cursor_data = data.get("cursor", {})
        total_cards = cursor_data.get("total", "?")

        log(f"  Стр. {page}: +{len(cards)} карточек "
            f"(просмотрено: {total_fetched}/{total_cards}), "
            f"совпадений: {len(results)}/{len(target_set)}")

        if len(results) >= len(target_set):
            log(f"  Все {len(target_set)} артикулов найдены.")
            break

        next_updated = cursor_data.get("updatedAt")
        next_nmid = cursor_data.get("nmID")
        if not next_updated and not next_nmid:
            break

        cursor_payload = {
            "limit": 100,
            "updatedAt": next_updated or "",
            "nmID": next_nmid or 0,
        }
        time.sleep(REQUEST_DELAY_SECONDS)

    elapsed = time.time() - start
    not_found = target_set - set(results.keys())
    log(f"Итого: {page} запросов за {elapsed:.1f} сек. "
        f"Найдено: {len(results)}/{len(target_set)}.")
    if not_found:
        log(f"Не найдено: {len(not_found)} артикулов (проставится 0 / пусто).")
        for nm in not_found:
            results[nm] = _empty_record()

    return results


# ---------------------------------------------------------------------
# Публичный API
# ---------------------------------------------------------------------

def cache_exists() -> bool:
    """
    Проверяет наличие файла photo_cache.xlsx. Используем os.path.exists
    со строкой пути (надёжнее Path.exists() на UNC + DFS).
    """
    return os.path.exists(str(config.PHOTO_CACHE_FILE))


def get_photo_data(target_nm_ids, use_cache: bool = False, log=print) -> dict:
    """
    Основная точка входа.
    Возвращает {str(nmID) -> {'photos': int, 'created_at': str}}.

      use_cache=True  + кэш существует  → читаем из xlsx-кэша, без обращения к API.
      use_cache=True  + кэша НЕТ         → API + сохранение кэша (защита от ошибки).
      use_cache=False                    → API + сохранение кэша.
    """
    cache_path = config.PHOTO_CACHE_FILE

    if use_cache:
        if cache_exists():
            log(f"✓ Читаем фото-кэш: {cache_path}")
            logger.info(f"get_photo_data: используем кэш {cache_path}")
            try:
                return load_photo_cache()
            except Exception as exc:
                logger.exception("Кэш повреждён — падаем на API")
                log(f"⚠ Кэш повреждён ({exc}), запрашиваем через API.")
        else:
            log("=" * 60)
            log(f"⚠ ВНИМАНИЕ: галочка «Использовать кэш фото» включена,")
            log(f"   но файл кэша НЕ НАЙДЕН: {cache_path}")
            log(f"   Будет выполнен полный запрос к WB API (5–15 минут).")
            log(f"   После окончания кэш создастся автоматически — следующий")
            log(f"   запуск с галочкой будет быстрым.")
            log("=" * 60)
            logger.warning(f"use_cache=True, но кэш отсутствует: {cache_path}. "
                           f"Переключаемся на API.")

    log("Запрос данных карточек через WB API...")
    logger.info(f"get_photo_data: WB API для {len(target_nm_ids)} nmID")
    try:
        results = fetch_photos_from_api(target_nm_ids, log=log)
    except Exception:
        logger.exception("Ошибка fetch_photos_from_api")
        raise

    # Сохраняем кэш БЕЗУСЛОВНО после каждого обращения к API
    try:
        saved_to = save_photo_cache(results)
        log(f"✓ Фото-кэш сохранён: {saved_to} ({len(results)} записей)")
    except Exception as exc:
        log(f"⚠ ВНИМАНИЕ: не удалось сохранить фото-кэш: {exc}")
        logger.exception("Не удалось сохранить фото-кэш — продолжаем без него")

    return results


def get_photo_counts(target_nm_ids, use_cache: bool = False, log=print) -> dict:
    """
    Backward-compat обёртка: возвращает только {nmID: photos_count}.
    Используется кодом, которому не нужна дата создания.
    """
    data = get_photo_data(target_nm_ids, use_cache=use_cache, log=log)
    return {k: v.get('photos', 0) for k, v in data.items()}
