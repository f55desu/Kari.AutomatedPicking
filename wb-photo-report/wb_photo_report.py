# -*- coding: utf-8 -*-
"""
WB product card photo counter.

What it does:
1. Reads WB nmID articles from manual input or a text file.
2. Loads WB Content API token from a .env file.
3. Requests WB product cards by targeted nmID search.
4. Builds an Excel report with:
   - sheet "Найдено"    -> nmID and photo count
   - sheet "Не найдено" -> nmID values not found in cards
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable
from urllib import error, request

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    raise SystemExit(
        "Не найдена библиотека openpyxl. Установите ее командой: "
        "python -m pip install openpyxl"
    ) from exc

# sys.stdout.reconfigure(encoding="utf-8")

API_URL = "https://content-api.wildberries.ru/content/v2/get/cards/list"
DEFAULT_ENV_FILE = ".env"
DEFAULT_PAGE_LIMIT = 100
REQUEST_DELAY_SECONDS = 0.65
MAX_RETRIES = 4
TOKEN_ENV_NAMES = (
    "WB_API_TOKEN",
    "WB_CONTENT_API_TOKEN",
    "WB_TOKEN",
)
RETRYABLE_STATUS_CODES = {408, 429, 500, 502, 503, 504}


class WBPhotoReportError(Exception):
    """Raised when the report cannot be built."""


@dataclass(frozen=True)
class ReportRow:
    nm_id: str
    photo_count: int


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Создает Excel-отчет по количеству фото в карточках WB."
    )
    parser.add_argument(
        "--env-file",
        default=DEFAULT_ENV_FILE,
        help="Путь к .env файлу с токеном WB. По умолчанию: .env",
    )
    parser.add_argument(
        "--input-file",
        help="Необязательный txt-файл со списком nmID. Если не указан, список запрашивается вручную.",
    )
    parser.add_argument(
        "--output",
        help="Имя итогового xlsx-файла. По умолчанию создается timestamp-имя.",
    )
    parser.add_argument(
        "--page-limit",
        type=int,
        default=DEFAULT_PAGE_LIMIT,
        help="Размер страницы при обходе карточек WB. По умолчанию: 100.",
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=60,
        help="Таймаут HTTP-запроса в секундах. По умолчанию: 60.",
    )
    return parser.parse_args()


def load_env_file(env_path: Path) -> None:
    if not env_path.exists():
        return

    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key:
            os.environ.setdefault(key, value)


def get_api_token() -> str:
    for env_name in TOKEN_ENV_NAMES:
        token = os.environ.get(env_name, "").strip()
        if token:
            return token
    expected = ", ".join(TOKEN_ENV_NAMES)
    raise WBPhotoReportError(
        f"Не найден токен WB API. Добавьте его в .env в одну из переменных: {expected}"
    )


def read_nm_ids_from_file(file_path: Path) -> str:
    if not file_path.exists():
        raise WBPhotoReportError(f"Файл со списком nmID не найден: {file_path}")
    return file_path.read_text(encoding="utf-8")


def read_nm_ids_from_prompt() -> str:
    print("Вставьте список nmID WB. Завершите ввод пустой строкой.")
    lines: list[str] = []
    while True:
        try:
            line = input().strip()
        except EOFError:
            break
        if not line:
            if lines:
                break
            continue
        lines.append(line)
    return "\n".join(lines)


def parse_nm_ids(raw_text: str) -> list[str]:
    matches = re.findall(r"\d+", raw_text)
    unique_ids: list[str] = []
    seen: set[str] = set()
    for match in matches:
        normalized = str(int(match))
        if normalized not in seen:
            seen.add(normalized)
            unique_ids.append(normalized)
    if not unique_ids:
        raise WBPhotoReportError("Не удалось распознать ни одного nmID в переданном списке.")
    return unique_ids


def build_output_path(output_arg: str | None) -> Path:
    if output_arg:
        return Path(output_arg).resolve()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return Path(f"wb_photo_report_{timestamp}.xlsx").resolve()


def wb_post_json(token: str, payload: dict, timeout: int) -> dict:
    body = json.dumps(payload).encode("utf-8")
    req = request.Request(
        API_URL,
        data=body,
        headers={
            "Authorization": token,
            "Content-Type": "application/json",
            "Accept": "application/json",
        },
        method="POST",
    )

    last_error_message = "неизвестная ошибка"

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            with request.urlopen(req, timeout=timeout) as response:
                response_text = response.read().decode("utf-8")
        except error.HTTPError as exc:
            details = exc.read().decode("utf-8", errors="ignore")
            last_error_message = f"WB API вернул HTTP {exc.code}. Детали: {details or exc.reason}"
            if exc.code in RETRYABLE_STATUS_CODES and attempt < MAX_RETRIES:
                sleep_seconds = max(REQUEST_DELAY_SECONDS, attempt * 1.5)
                print(f"Повтор запроса после HTTP {exc.code}, попытка {attempt}/{MAX_RETRIES}...")
                time.sleep(sleep_seconds)
                continue
            raise WBPhotoReportError(last_error_message) from exc
        except error.URLError as exc:
            last_error_message = f"Не удалось обратиться к WB API: {exc.reason}"
            if attempt < MAX_RETRIES:
                sleep_seconds = max(REQUEST_DELAY_SECONDS, attempt * 1.5)
                print(f"Повтор запроса после сетевой ошибки, попытка {attempt}/{MAX_RETRIES}...")
                time.sleep(sleep_seconds)
                continue
            raise WBPhotoReportError(last_error_message) from exc

        try:
            parsed = json.loads(response_text)
        except json.JSONDecodeError as exc:
            raise WBPhotoReportError(
                "WB API вернул ответ, который не удалось разобрать как JSON."
            ) from exc

        if parsed.get("error") is True:
            error_text = parsed.get("errorText") or "без текста ошибки"
            last_error_message = f"WB API сообщил об ошибке: {error_text}"
            if "timeout" in error_text.lower() and attempt < MAX_RETRIES:
                sleep_seconds = max(REQUEST_DELAY_SECONDS, attempt * 1.5)
                print(f"Повтор запроса после timeout от WB, попытка {attempt}/{MAX_RETRIES}...")
                time.sleep(sleep_seconds)
                continue
            raise WBPhotoReportError(last_error_message)

        return parsed

    raise WBPhotoReportError(last_error_message)


def build_cards_payload(limit: int, nm_id: str) -> dict:
    return {
        "settings": {
            "cursor": {
                "limit": limit,
            },
            "filter": {
                "textSearch": nm_id,
                "withPhoto": -1,
            },
        }
    }


def count_photos(card: dict) -> int:
    photos = card.get("photos")
    if isinstance(photos, list):
        return len(photos)
    return 0


def find_card_by_nm_id(token: str, nm_id: str, page_limit: int, timeout: int) -> ReportRow | None:
    payload = build_cards_payload(limit=page_limit, nm_id=nm_id)
    response = wb_post_json(token=token, payload=payload, timeout=timeout)
    cards = response.get("cards") or []

    for card in cards:
        card_nm_id = str(card.get("nmID", "")).strip()
        if card_nm_id == nm_id:
            return ReportRow(nm_id=nm_id, photo_count=count_photos(card))
    return None


def collect_photo_counts(
    token: str,
    requested_nm_ids: list[str],
    page_limit: int,
    timeout: int,
) -> tuple[list[ReportRow], list[str], int]:
    found_rows: list[ReportRow] = []
    not_found_rows: list[str] = []

    total_requested = len(requested_nm_ids)
    for index, nm_id in enumerate(requested_nm_ids, start=1):
        print(f"[{index}/{total_requested}] Проверяю nmID {nm_id}...")
        row = find_card_by_nm_id(
            token=token,
            nm_id=nm_id,
            page_limit=page_limit,
            timeout=timeout,
        )

        if row is None:
            not_found_rows.append(nm_id)
            print(f"[{index}/{total_requested}] nmID {nm_id}: не найден")
        else:
            found_rows.append(row)
            print(
                f"[{index}/{total_requested}] nmID {nm_id}: найдено фото {row.photo_count}"
            )

        if index < total_requested:
            time.sleep(REQUEST_DELAY_SECONDS)

    return found_rows, not_found_rows, total_requested


def style_header(cell) -> None:
    cell.font = Font(bold=True, color="FFFFFF", name="Calibri")
    cell.fill = PatternFill("solid", fgColor="1F4E78")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    side = Side(style="thin", color="BFBFBF")
    cell.border = Border(left=side, right=side, top=side, bottom=side)


def style_data_cell(cell) -> None:
    cell.alignment = Alignment(horizontal="center", vertical="center")
    side = Side(style="thin", color="D9D9D9")
    cell.border = Border(left=side, right=side, top=side, bottom=side)


def autofit_columns(ws) -> None:
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[column_letter].width = min(max_length + 4, 40)


def write_rows(ws, headers: Iterable[str], rows: Iterable[Iterable[object]]) -> None:
    headers = list(headers)
    rows = list(rows)

    ws.append(headers)
    for cell in ws[1]:
        style_header(cell)

    for row in rows:
        ws.append(list(row))

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            style_data_cell(cell)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autofit_columns(ws)


def build_excel_report(
    found_rows: list[ReportRow],
    not_found_rows: list[str],
    output_path: Path,
) -> None:
    wb = openpyxl.Workbook()
    ws_found = wb.active
    ws_found.title = "Найдено"
    write_rows(
        ws_found,
        headers=("Артикул WB", "Количество фото"),
        rows=((row.nm_id, row.photo_count) for row in found_rows),
    )

    ws_missing = wb.create_sheet("Не найдено")
    write_rows(
        ws_missing,
        headers=("Артикул WB",),
        rows=((nm_id,) for nm_id in not_found_rows),
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> int:
    args = parse_args()
    env_path = Path(args.env_file).resolve()
    load_env_file(env_path)

    token = get_api_token()
    raw_nm_ids = (
        read_nm_ids_from_file(Path(args.input_file).resolve())
        if args.input_file
        else read_nm_ids_from_prompt()
    )
    nm_ids = parse_nm_ids(raw_nm_ids)

    if args.page_limit < 1:
        raise WBPhotoReportError("--page-limit должен быть положительным числом.")

    output_path = build_output_path(args.output)
    print(f"Всего уникальных nmID в запросе: {len(nm_ids)}")
    print("Запрашиваю карточки WB и считаю фотографии...")

    found_rows, not_found_rows, pages_scanned = collect_photo_counts(
        token=token,
        requested_nm_ids=nm_ids,
        page_limit=args.page_limit,
        timeout=args.timeout,
    )

    build_excel_report(
        found_rows=found_rows,
        not_found_rows=not_found_rows,
        output_path=output_path,
    )

    print("")
    print(f"Готово. Отчет сохранен: {output_path}")
    print(f"Выполнено точечных запросов: {pages_scanned}")
    print(f"Найдено артикулов: {len(found_rows)}")
    print(f"Не найдено артикулов: {len(not_found_rows)}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except KeyboardInterrupt:
        print("\nОперация прервана пользователем.")
        raise SystemExit(130)
    except WBPhotoReportError as exc:
        print(f"Ошибка: {exc}")
        raise SystemExit(1)
